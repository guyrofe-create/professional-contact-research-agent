from __future__ import annotations

import json
import shutil
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
import agent

ALGO_VERSION = agent.ALGO_VERSION
OUT = Path('output')
FINAL = Path('final')


def xml_safe(value):
    if not isinstance(value, str):
        return value
    # Keep XML 1.0 legal characters only. This explicitly removes NUL and
    # other control characters regardless of pandas string/object dtype.
    return ''.join(
        ch for ch in value
        if ch in ('\t', '\n', '\r')
        or 0x20 <= ord(ch) <= 0xD7FF
        or 0xE000 <= ord(ch) <= 0xFFFD
        or 0x10000 <= ord(ch) <= 0x10FFFF
    )


def sanitize_frame(frame: pd.DataFrame) -> pd.DataFrame:
    safe = frame.copy()
    for col in safe.columns:
        safe[col] = safe[col].map(xml_safe)
    return safe


def read_checkpoint() -> pd.DataFrame:
    cp = OUT / 'checkpoint.jsonl'
    if not cp.exists():
        raise SystemExit('Missing output/checkpoint.jsonl')
    done = {}
    for line in cp.read_text(encoding='utf-8', errors='ignore').splitlines():
        try:
            row = json.loads(line)
        except Exception:
            continue
        if row.get('algo_version') == ALGO_VERSION:
            done[(row.get('name', ''), row.get('category', ''))] = row
    if not done:
        raise SystemExit(f'No version-{ALGO_VERSION} checkpoint rows found')
    return pd.DataFrame(list(done.values()))


def validate_xlsx(path: Path, expected_min_rows: int = 0):
    if not path.exists() or path.stat().st_size == 0:
        raise SystemExit(f'Missing/empty workbook: {path}')
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = max(0, ws.max_row - 1)
    wb.close()
    if rows < expected_min_rows:
        raise SystemExit(f'Workbook row regression: {path} has {rows}, expected at least {expected_min_rows}')
    print(f'VALIDATED {path} rows={rows} bytes={path.stat().st_size}')


def main():
    OUT.mkdir(exist_ok=True)
    FINAL.mkdir(exist_ok=True)

    frame = read_checkpoint()
    frame = frame.sort_values(['priority', 'status', 'confidence'], ascending=[True, True, False])
    frame = frame.drop_duplicates(subset=['name', 'category'], keep='first')

    expanded = agent.annotate_shared_contacts(agent.expand_verified_contacts(frame))

    safe_frame = sanitize_frame(frame)
    safe_frame.to_csv(OUT / 'audit.csv', index=False, encoding='utf-8-sig')
    safe_frame.to_excel(OUT / 'audit.xlsx', index=False)

    found = expanded[expanded.send_eligible].copy() if not expanded.empty else pd.DataFrame(columns=list(frame.columns) + ['candidate_rank'])
    found = found.sort_values(['priority', 'confidence'], ascending=[True, False]).drop_duplicates(subset=['email'], keep='first')
    safe_found = sanitize_frame(found)
    safe_found.to_csv(OUT / 'contacts.csv', index=False, encoding='utf-8-sig')
    safe_found.to_excel(OUT / 'contacts.xlsx', index=False)

    personal = found[found.outreach_scope == 'PERSON'].copy() if not found.empty else found.copy()
    organization = found[found.outreach_scope == 'ORGANIZATION_OR_SHARED_ROUTE'].copy() if not found.empty else found.copy()
    shared = found[found.shared_contact == True].copy() if not found.empty else found.copy()
    for name, partition in (('personal', personal), ('organization', organization), ('shared', shared)):
        safe_partition = sanitize_frame(partition)
        safe_partition.to_csv(OUT / f'contacts_{name}.csv', index=False, encoding='utf-8-sig')
        safe_partition.to_excel(OUT / f'contacts_{name}.xlsx', index=False)

    review = sanitize_frame(frame[frame.status.str.startswith('REVIEW')].copy())
    review.to_excel(OUT / 'review.xlsx', index=False)

    target_total = len(pd.read_csv('targets.csv')) if Path('targets.csv').exists() else len(frame)
    fanout = expanded.groupby('email').size() if not expanded.empty else pd.Series(dtype=int)
    summary = {
        'algo_version': ALGO_VERSION,
        'total_targets': int(target_total),
        'touched_targets': int(len(frame)),
        'resolved_targets': int((~frame.status.str.startswith('PENDING')).sum()),
        'verified': int((frame.status == 'VERIFIED').sum()),
        'not_verified': int((frame.status == 'NO_VERIFIED_PUBLIC_EMAIL').sum()),
        'pending': int(frame.status.str.startswith('PENDING').sum()),
        'review': int(frame.status.str.startswith('REVIEW').sum()),
        'unique_emails': int(found.email.nunique()),
        'personal_unique_emails': int(personal.email.nunique()) if not personal.empty else 0,
        'organization_unique_emails': int(organization.email.nunique()) if not organization.empty else 0,
        'shared_unique_emails': int(shared.email.nunique()) if not shared.empty else 0,
        'max_email_target_fanout': int(fanout.max()) if not fanout.empty else 0,
        'emails_with_fanout_over_5': int((fanout > 5).sum()) if not fanout.empty else 0,
        'personalization_safe_emails': int(found.personalization_safe.sum()) if not found.empty else 0,
        'organization_or_shared_routes': int((found.outreach_scope == 'ORGANIZATION_OR_SHARED_ROUTE').sum()) if not found.empty else 0,
        'direct_emails': int(found.extraction_method.fillna('').str.startswith(('direct_', 'official_')).sum()) if not found.empty else 0,
        'institutional_emails': int((found.email_type == 'CLINIC_OR_ORGANIZATION').sum()) if not found.empty else 0,
        'by_category': frame.groupby('category').status.value_counts().unstack(fill_value=0).to_dict('index'),
    }
    (OUT / 'summary.json').write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding='utf-8')

    # Validate the generated workbooks before promoting them to FINAL.
    validate_xlsx(OUT / 'audit.xlsx', expected_min_rows=len(frame))
    validate_xlsx(OUT / 'contacts.xlsx', expected_min_rows=len(found))
    validate_xlsx(OUT / 'contacts_personal.xlsx', expected_min_rows=len(personal))
    validate_xlsx(OUT / 'contacts_organization.xlsx', expected_min_rows=len(organization))
    validate_xlsx(OUT / 'contacts_shared.xlsx', expected_min_rows=len(shared))
    validate_xlsx(OUT / 'review.xlsx', expected_min_rows=len(review))

    completion = (OUT / 'COMPLETE.txt').read_text(encoding='utf-8', errors='ignore') if (OUT / 'COMPLETE.txt').exists() else ''
    if completion.startswith('COMPLETE') and f'algo_version={ALGO_VERSION}' in completion:
        shutil.copy2(OUT / 'contacts.xlsx', FINAL / 'contacts_FINAL.xlsx')
        shutil.copy2(OUT / 'contacts_personal.xlsx', FINAL / 'contacts_personal_FINAL.xlsx')
        shutil.copy2(OUT / 'contacts_organization.xlsx', FINAL / 'contacts_organization_FINAL.xlsx')
        shutil.copy2(OUT / 'contacts_shared.xlsx', FINAL / 'contacts_shared_FINAL.xlsx')
        shutil.copy2(OUT / 'audit.xlsx', FINAL / 'audit_FINAL.xlsx')
        shutil.copy2(OUT / 'review.xlsx', FINAL / 'review_FINAL.xlsx')
        shutil.copy2(OUT / 'summary.json', FINAL / 'summary_FINAL.json')
        shutil.copy2(OUT / 'COMPLETE.txt', FINAL / 'COMPLETE.txt')
        if Path('seed_summary.json').exists():
            shutil.copy2('seed_summary.json', FINAL / 'seed_summary_FINAL.json')
        validate_xlsx(FINAL / 'contacts_FINAL.xlsx', expected_min_rows=len(found))
        validate_xlsx(FINAL / 'audit_FINAL.xlsx', expected_min_rows=len(frame))
        validate_xlsx(FINAL / 'review_FINAL.xlsx', expected_min_rows=len(review))
    else:
        print('Research incomplete; current reports validated but FINAL files were not replaced.')
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == '__main__':
    main()
